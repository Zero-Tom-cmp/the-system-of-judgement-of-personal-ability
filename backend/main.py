import csv
import io
import json
import os
from datetime import datetime, timedelta, timezone

import bcrypt
import jwt
from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel

from database import get_db, init_db

def audit_log(cursor, operator, action, target_type, target_id, detail=""):
    cursor.execute(
        "INSERT INTO audit_log (operator, action, target_type, target_id, detail) VALUES (?, ?, ?, ?, ?)",
        (operator, action, target_type, str(target_id), detail),
    )


def _save_conflicts_to_disk(sheet_name, errors):
    """将冲突数据按学号分组，每个学号一个CSV，放入时间戳文件夹"""
    import os, time as _time
    conflicts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "conflicts")
    os.makedirs(conflicts_dir, exist_ok=True)
    timestamp = _time.strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(conflicts_dir, f"conflicts_{sheet_name}_{timestamp}")
    os.makedirs(folder, exist_ok=True)

    # 按学号分组
    groups = {}
    for e in errors:
        if isinstance(e, dict) and "data" in e:
            sid = e["data"].get("学号", "unknown")
            if sid not in groups:
                groups[sid] = []
            groups[sid].append(e)

    if not groups:
        return folder

    headers = list(errors[0]["data"].keys()) if (errors and isinstance(errors[0], dict)) else []
    for sid, items in groups.items():
        filepath = os.path.join(folder, f"{sid}.csv")
        with open(filepath, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["行号", "失败原因"] + headers)
            for e in items:
                writer.writerow([e.get("row", ""), e.get("reason", "")] + [e["data"].get(h, "") for h in headers])

    # 写一个汇总文件
    summary_path = os.path.join(folder, "_汇总.csv")
    with open(summary_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["学号", "冲突条数"])
        for sid, items in sorted(groups.items()):
            writer.writerow([sid, len(items)])

    return folder
from evaluation import evaluate_student_abilities, match_jobs

app = FastAPI(title="人岗匹配评估系统")

# 简易登录限流
_login_attempts = {}  # {ip: (count, first_attempt_time)}


@app.get("/api/health")
def health_check():
    try:
        conn = get_db()
        conn.execute("SELECT 1")
        conn.close()
        return {"status": "ok", "database": "connected"}
    except Exception as e:
        return {"status": "error", "database": str(e)}


@app.middleware("http")
async def rate_limit_login(request: Request, call_next):
    if request.url.path == "/api/login" and request.method == "POST":
        ip = request.client.host if request.client else "unknown"
        now = datetime.now(timezone.utc)
        if ip in _login_attempts:
            count, first = _login_attempts[ip]
            if (now - first).seconds > 300:  # 5分钟窗口
                _login_attempts[ip] = (1, now)
            elif count >= 20:  # 5分钟内最多20次
                return JSONResponse(status_code=429, content={"detail": "登录尝试过于频繁，请5分钟后再试"})
            else:
                _login_attempts[ip] = (count + 1, first)
        else:
            _login_attempts[ip] = (1, now)
    response = await call_next(request)
    return response

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SECRET_KEY = os.getenv("JWT_SECRET", "person-job-match-secret-key-change-in-production")
ALGORITHM = "HS256"


class LoginRequest(BaseModel):
    username: str
    password: str


def create_token(student_id: str, role: str) -> str:
    payload = {
        "student_id": student_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=12),
    }
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)


def verify_token(token: str) -> dict:
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="无效的认证令牌")


def require_admin(payload: dict):
    if payload.get("role") != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")


@app.on_event("startup")
def startup():
    init_db()


@app.post("/api/login")
def login(req: LoginRequest):
    conn = get_db()
    try:
        user = conn.execute(
            "SELECT * FROM students WHERE student_id = ?", (req.username,)
        ).fetchone()

        if not user:
            raise HTTPException(status_code=401, detail="用户名或密码错误")

        if not bcrypt.checkpw(req.password.encode(), user["password_hash"].encode()):
            raise HTTPException(status_code=401, detail="用户名或密码错误")

        token = create_token(user["student_id"], user["role"])
        return {
            "token": token,
            "student_id": user["student_id"],
            "name": user["name"],
            "role": user["role"],
            "major": user["major"],
        }
    finally:
        conn.close()


@app.get("/api/student/{student_id}/info")
def get_student_info(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if payload["role"] != "admin" and payload["student_id"] != student_id:
        raise HTTPException(status_code=403, detail="只能查看自己的信息")

    conn = get_db()
    try:
        user = conn.execute(
            "SELECT student_id, name, college, major, class_name, gpa, role FROM students WHERE student_id = ?",
            (student_id,),
        ).fetchone()
        if not user:
            raise HTTPException(status_code=404, detail="学生不存在")

        result = dict(user)
        # 计算已修学分
        credits_row = conn.execute(
            "SELECT SUM(credit) as total FROM courses WHERE student_id = ?",
            (student_id,),
        ).fetchone()
        result["total_credits"] = round(credits_row["total"], 1) if credits_row["total"] else 0
        return result
    finally:
        conn.close()


@app.get("/api/student/{student_id}/courses")
def get_student_courses(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if payload["role"] != "admin" and payload["student_id"] != student_id:
        raise HTTPException(status_code=403, detail="只能查看自己的信息")

    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM courses WHERE student_id = ? ORDER BY semester",
            (student_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.get("/api/student/{student_id}/competitions")
def get_student_competitions(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if payload["role"] != "admin" and payload["student_id"] != student_id:
        raise HTTPException(status_code=403, detail="只能查看自己的信息")

    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM competitions WHERE student_id = ? ORDER BY date DESC",
            (student_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.get("/api/student/{student_id}/internships")
def get_student_internships(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if payload["role"] != "admin" and payload["student_id"] != student_id:
        raise HTTPException(status_code=403, detail="只能查看自己的信息")

    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM internships WHERE student_id = ? ORDER BY period DESC",
            (student_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.get("/api/student/{student_id}/projects")
def get_student_projects(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if payload["role"] != "admin" and payload["student_id"] != student_id:
        raise HTTPException(status_code=403, detail="只能查看自己的信息")

    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM projects WHERE student_id = ? ORDER BY period DESC",
            (student_id,),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.get("/api/student/{student_id}/abilities")
def get_student_abilities(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if payload["role"] != "admin" and payload["student_id"] != student_id:
        raise HTTPException(status_code=403, detail="只能查看自己的信息")

    conn = get_db()
    try:
        return evaluate_student_abilities(conn.cursor(), student_id)
    finally:
        conn.close()


@app.get("/api/student/{student_id}/job-match")
def get_job_match(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    if payload["role"] != "admin" and payload["student_id"] != student_id:
        raise HTTPException(status_code=403, detail="只能查看自己的信息")

    conn = get_db()
    try:
        return match_jobs(conn.cursor(), student_id)
    finally:
        conn.close()


@app.get("/api/admin/students")
def admin_list_students(request: Request, page: int = 1, page_size: int = 100, search: str = "", college: str = "", major: str = ""):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)

    conn = get_db()
    try:
        where = ["role = 'student'"]
        params = []
        if search:
            where.append("(student_id LIKE ? OR name LIKE ?)")
            params.extend([f"%{search}%", f"%{search}%"])
        if college:
            where.append("college = ?")
            params.append(college)
        if major:
            where.append("major = ?")
            params.append(major)

        where_clause = " AND ".join(where)
        count = conn.execute(f"SELECT COUNT(*) as c FROM students WHERE {where_clause}", params).fetchone()["c"]

        offset = (page - 1) * page_size
        rows = conn.execute(
            f"SELECT student_id, name, college, major, class_name, gpa FROM students WHERE {where_clause} ORDER BY student_id LIMIT ? OFFSET ?",
            params + [page_size, offset],
        ).fetchall()
        return {"data": [dict(r) for r in rows], "total": count, "page": page, "page_size": page_size}
    finally:
        conn.close()


@app.get("/api/admin/search")
def admin_search(q: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)

    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT student_id, name, college, major, class_name, gpa FROM students WHERE role = 'student' AND (student_id LIKE ? OR name LIKE ?)",
            (f"%{q}%", f"%{q}%"),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.get("/api/admin/stats")
def admin_stats(request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        total = conn.execute("SELECT COUNT(*) as c FROM students WHERE role='student'").fetchone()["c"]
        majors = conn.execute(
            "SELECT major, COUNT(*) as c FROM students WHERE role='student' GROUP BY major"
        ).fetchall()
        avg_gpa = conn.execute(
            "SELECT AVG(gpa) as avg FROM students WHERE role='student'"
        ).fetchone()["avg"]
        total_courses = conn.execute("SELECT COUNT(*) as c FROM courses").fetchone()["c"]
        total_comps = conn.execute("SELECT COUNT(*) as c FROM competitions").fetchone()["c"]
        total_interns = conn.execute("SELECT COUNT(*) as c FROM internships").fetchone()["c"]
        total_projects = conn.execute("SELECT COUNT(*) as c FROM projects").fetchone()["c"]
        return {
            "total_students": total,
            "majors": [{"major": r["major"], "count": r["c"]} for r in majors],
            "avg_gpa": round(avg_gpa, 2) if avg_gpa else 0,
            "total_courses": total_courses,
            "total_competitions": total_comps,
            "total_internships": total_interns,
            "total_projects": total_projects,
        }
    finally:
        conn.close()


# ========== 管理员：学生 CRUD ==========

class StudentCreate(BaseModel):
    student_id: str
    name: str
    college: str
    major: str
    class_name: str
    gpa: float
    password: str

class StudentUpdate(BaseModel):
    name: str = None
    college: str = None
    major: str = None
    class_name: str = None
    gpa: float = None
    password: str = None


@app.post("/api/admin/student")
def admin_create_student(req: StudentCreate, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        existing = conn.execute("SELECT id FROM students WHERE student_id = ?", (req.student_id,)).fetchone()
        if existing:
            raise HTTPException(status_code=400, detail="学号已存在")
        pwd_hash = bcrypt.hashpw(req.password.encode(), bcrypt.gensalt()).decode()
        conn.execute(
            "INSERT INTO students (student_id, name, college, major, class_name, gpa, password_hash, role) VALUES (?, ?, ?, ?, ?, ?, ?, 'student')",
            (req.student_id, req.name, req.college, req.major, req.class_name, req.gpa, pwd_hash),
        )
        audit_log(conn.cursor(), payload["student_id"], "创建", "student", req.student_id)
        conn.commit()
        return {"message": "创建成功", "student_id": req.student_id}
    finally:
        conn.close()


@app.put("/api/admin/student/{student_id}")
def admin_update_student(student_id: str, req: StudentUpdate, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        existing = conn.execute("SELECT * FROM students WHERE student_id = ?", (student_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="学生不存在")
        updates = []
        params = []
        if req.name is not None:
            updates.append("name = ?"); params.append(req.name)
        if req.college is not None:
            updates.append("college = ?"); params.append(req.college)
        if req.major is not None:
            updates.append("major = ?"); params.append(req.major)
        if req.class_name is not None:
            updates.append("class_name = ?"); params.append(req.class_name)
        if req.gpa is not None:
            updates.append("gpa = ?"); params.append(req.gpa)
        if req.password is not None:
            pwd_hash = bcrypt.hashpw(req.password.encode(), bcrypt.gensalt()).decode()
            updates.append("password_hash = ?"); params.append(pwd_hash)
        if updates:
            params.append(student_id)
            conn.execute(f"UPDATE students SET {', '.join(updates)} WHERE student_id = ?", params)
            conn.commit()
        return {"message": "更新成功"}
    finally:
        conn.close()


@app.delete("/api/admin/student/{student_id}")
def admin_delete_student(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        existing = conn.execute("SELECT * FROM students WHERE student_id = ? AND role='student'", (student_id,)).fetchone()
        if not existing:
            raise HTTPException(status_code=404, detail="学生不存在")
        for table in ["courses", "competitions", "internships", "projects"]:
            conn.execute(f"DELETE FROM {table} WHERE student_id = ?", (student_id,))
        conn.execute("DELETE FROM students WHERE student_id = ?", (student_id,))
        audit_log(conn.cursor(), payload["student_id"], "删除", "student", student_id)
        conn.commit()
        return {"message": "删除成功"}
    finally:
        conn.close()


# ========== 管理员：数据项 CRUD ==========

@app.post("/api/admin/student/{student_id}/courses")
async def admin_add_course(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    body = await request.json()
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO courses (student_id, course_name, score, credit, practical_credit, semester, course_nature) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (student_id, body["course_name"], body["score"], body["credit"], body.get("practical_credit", 0), body["semester"], body["course_nature"]),
        )
        conn.commit()
        return {"message": "添加成功"}
    finally:
        conn.close()


@app.put("/api/admin/course/{course_id}")
async def admin_update_course(course_id: int, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    body = await request.json()
    conn = get_db()
    try:
        updates = [f"{k} = ?" for k in ["course_name", "score", "credit", "practical_credit", "semester", "course_nature"] if k in body]
        params = [body[k] for k in ["course_name", "score", "credit", "practical_credit", "semester", "course_nature"] if k in body]
        if updates:
            params.append(course_id)
            conn.execute(f"UPDATE courses SET {', '.join(updates)} WHERE id = ?", params)
            conn.commit()
        return {"message": "更新成功"}
    finally:
        conn.close()


@app.delete("/api/admin/course/{course_id}")
def admin_delete_course(course_id: int, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        conn.execute("DELETE FROM courses WHERE id = ?", (course_id,))
        conn.commit()
        return {"message": "删除成功"}
    finally:
        conn.close()


# 竞赛
@app.post("/api/admin/student/{student_id}/competitions")
async def admin_add_competition(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    body = await request.json()
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO competitions (student_id, comp_name, level, award, rank, date) VALUES (?, ?, ?, ?, ?, ?)",
            (student_id, body["comp_name"], body["level"], body["award"], body["rank"], body["date"]),
        )
        conn.commit()
        return {"message": "添加成功"}
    finally:
        conn.close()


@app.put("/api/admin/competition/{comp_id}")
async def admin_update_competition(comp_id: int, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    body = await request.json()
    conn = get_db()
    try:
        cols = ["comp_name", "level", "award", "rank", "date"]
        updates = [f"{k} = ?" for k in cols if k in body]
        params = [body[k] for k in cols if k in body]
        if updates:
            params.append(comp_id)
            conn.execute(f"UPDATE competitions SET {', '.join(updates)} WHERE id = ?", params)
            conn.commit()
        return {"message": "更新成功"}
    finally:
        conn.close()


@app.delete("/api/admin/competition/{comp_id}")
def admin_delete_competition(comp_id: int, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        conn.execute("DELETE FROM competitions WHERE id = ?", (comp_id,))
        conn.commit()
        return {"message": "删除成功"}
    finally:
        conn.close()


# 企业实习
@app.post("/api/admin/student/{student_id}/internships")
async def admin_add_internship(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    body = await request.json()
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO internships (student_id, company, position, description, period) VALUES (?, ?, ?, ?, ?)",
            (student_id, body["company"], body["position"], body["description"], body["period"]),
        )
        conn.commit()
        return {"message": "添加成功"}
    finally:
        conn.close()


@app.put("/api/admin/internship/{intern_id}")
async def admin_update_internship(intern_id: int, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    body = await request.json()
    conn = get_db()
    try:
        cols = ["company", "position", "description", "period"]
        updates = [f"{k} = ?" for k in cols if k in body]
        params = [body[k] for k in cols if k in body]
        if updates:
            params.append(intern_id)
            conn.execute(f"UPDATE internships SET {', '.join(updates)} WHERE id = ?", params)
            conn.commit()
        return {"message": "更新成功"}
    finally:
        conn.close()


@app.delete("/api/admin/internship/{intern_id}")
def admin_delete_internship(intern_id: int, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        conn.execute("DELETE FROM internships WHERE id = ?", (intern_id,))
        conn.commit()
        return {"message": "删除成功"}
    finally:
        conn.close()


# 项目经历
@app.post("/api/admin/student/{student_id}/projects")
async def admin_add_project(student_id: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    body = await request.json()
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO projects (student_id, project_name, rank, description, period) VALUES (?, ?, ?, ?, ?)",
            (student_id, body["project_name"], body["rank"], body["description"], body["period"]),
        )
        conn.commit()
        return {"message": "添加成功"}
    finally:
        conn.close()


@app.put("/api/admin/project/{proj_id}")
async def admin_update_project(proj_id: int, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    body = await request.json()
    conn = get_db()
    try:
        cols = ["project_name", "rank", "description", "period"]
        updates = [f"{k} = ?" for k in cols if k in body]
        params = [body[k] for k in cols if k in body]
        if updates:
            params.append(proj_id)
            conn.execute(f"UPDATE projects SET {', '.join(updates)} WHERE id = ?", params)
            conn.commit()
        return {"message": "更新成功"}
    finally:
        conn.close()


@app.delete("/api/admin/project/{proj_id}")
def admin_delete_project(proj_id: int, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        conn.execute("DELETE FROM projects WHERE id = ?", (proj_id,))
        conn.commit()
        return {"message": "删除成功"}
    finally:
        conn.close()


@app.get("/api/training-plan/{major}")
def get_training_plan(major: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT plan_data FROM training_plan WHERE major = ?", (major,)
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="未找到该专业的培养方案")
        return json.loads(row["plan_data"])
    finally:
        conn.close()


# ========== 密码修改 ==========

class PasswordChange(BaseModel):
    old_password: str
    new_password: str


@app.put("/api/student/password")
def change_password(req: PasswordChange, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    student_id = payload["student_id"]

    conn = get_db()
    try:
        user = conn.execute("SELECT * FROM students WHERE student_id = ?", (student_id,)).fetchone()
        if not bcrypt.checkpw(req.old_password.encode(), user["password_hash"].encode()):
            raise HTTPException(status_code=400, detail="原密码错误")
        new_hash = bcrypt.hashpw(req.new_password.encode(), bcrypt.gensalt()).decode()
        conn.execute("UPDATE students SET password_hash = ? WHERE student_id = ?", (new_hash, student_id))
        audit_log(conn.cursor(), student_id, "修改密码", "student", student_id)
        conn.commit()
        return {"message": "密码修改成功"}
    finally:
        conn.close()


# ========== 审计日志 ==========

@app.get("/api/admin/audit-log")
def admin_audit_log(request: Request, limit: int = 50):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT * FROM audit_log ORDER BY created_at DESC LIMIT ?", (limit,)
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# ========== 评估规则配置 ==========

@app.get("/api/admin/rules/{rule_type}")
def admin_get_rules(rule_type: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        if rule_type == "ability_config":
            rows = conn.execute("SELECT * FROM ability_config ORDER BY major, id").fetchall()
        elif rule_type == "course_mapping":
            rows = conn.execute("SELECT * FROM course_ability_mapping ORDER BY course_name").fetchall()
        elif rule_type == "job_profile":
            rows = conn.execute("SELECT * FROM job_ability_profile ORDER BY job_name").fetchall()
        else:
            raise HTTPException(status_code=400, detail="无效的规则类型")
        return [dict(r) for r in rows]
    finally:
        conn.close()


@app.put("/api/admin/rules/{rule_type}/{rule_id}")
async def admin_update_rule(rule_type: str, rule_id: int, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    body = await request.json()
    conn = get_db()
    try:
        table_map = {
            "ability_config": "ability_config",
            "course_mapping": "course_ability_mapping",
            "job_profile": "job_ability_profile",
        }
        table = table_map.get(rule_type)
        if not table:
            raise HTTPException(status_code=400, detail="无效的规则类型")
        sets = ", ".join([f"{k} = ?" for k in body.keys()])
        params = list(body.values()) + [rule_id]
        conn.execute(f"UPDATE {table} SET {sets} WHERE id = ?", params)
        audit_log(conn.cursor(), payload["student_id"], "修改规则", rule_type, str(rule_id), str(body))
        conn.commit()
        return {"message": "更新成功"}
    finally:
        conn.close()


# ========== 批量操作 ==========

class BatchDeleteRequest(BaseModel):
    student_ids: list

class BatchResetPwdRequest(BaseModel):
    student_ids: list
    new_password: str = "student123"


@app.post("/api/admin/students/batch-delete")
def admin_batch_delete(req: BatchDeleteRequest, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        deleted = 0
        for sid in req.student_ids:
            if sid == "admin": continue
            existing = conn.execute("SELECT id FROM students WHERE student_id = ? AND role='student'", (sid,)).fetchone()
            if existing:
                for table in ["courses", "competitions", "internships", "projects"]:
                    conn.execute(f"DELETE FROM {table} WHERE student_id = ?", (sid,))
                conn.execute("DELETE FROM students WHERE student_id = ?", (sid,))
                audit_log(conn.cursor(), payload["student_id"], "批量删除", "student", sid)
                deleted += 1
        conn.commit()
        return {"message": f"成功删除 {deleted} 名学生", "deleted": deleted}
    finally:
        conn.close()


@app.post("/api/admin/students/batch-reset-pwd")
def admin_batch_reset_pwd(req: BatchResetPwdRequest, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    conn = get_db()
    try:
        pwd_hash = bcrypt.hashpw(req.new_password.encode(), bcrypt.gensalt()).decode()
        count = 0
        for sid in req.student_ids:
            conn.execute("UPDATE students SET password_hash = ? WHERE student_id = ? AND role='student'", (pwd_hash, sid))
            audit_log(conn.cursor(), payload["student_id"], "重置密码", "student", sid)
            count += 1
        conn.commit()
        return {"message": f"成功重置 {count} 名学生的密码", "count": count}
    finally:
        conn.close()


def _log_import_error(error_type, detail, file_name=""):
    """将导入失败的错误原因写入日志文件"""
    import os, time as _time, traceback
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "conflicts", "import_errors")
    os.makedirs(log_dir, exist_ok=True)
    timestamp = _time.strftime("%Y%m%d_%H%M%S")
    filepath = os.path.join(log_dir, f"error_{timestamp}.log")
    with open(filepath, "w", encoding="utf-8") as f:
        f.write(f"时间: {_time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"类型: {error_type}\n")
        f.write(f"文件: {file_name}\n")
        f.write(f"详情: {detail}\n")
        if isinstance(detail, Exception):
            f.write(f"堆栈:\n{traceback.format_exc()}\n")
    return filepath


# ========== 综合导入（Excel，需注册在批量导入之前） ==========

SHEET_CONFIG = {
    "学生基本信息": {
        "table": "students", "columns": ["student_id", "name", "college", "major", "class_name", "gpa", "password_hash"],
        "headers": ["学号", "姓名", "学院", "专业", "班级", "GPA", "密码"], "mode": "upsert",
        "conflict_keys": None,  # upsert 模式不检测冲突
    },
    "课程成绩": {
        "table": "courses", "columns": ["student_id", "course_name", "score", "credit", "practical_credit", "semester", "course_nature"],
        "headers": ["学号", "课程名称", "成绩", "学分", "实践学分", "学期", "类别"], "mode": "insert",
        "conflict_keys": ["student_id", "course_name"],
        "conflict_sql": "SELECT id FROM courses WHERE student_id = ? AND course_name = ?",
    },
    "竞赛获奖": {
        "table": "competitions", "columns": ["student_id", "comp_name", "level", "award", "rank", "date"],
        "headers": ["学号", "竞赛名称", "级别", "获奖", "排名", "时间"], "mode": "insert",
        "conflict_keys": ["student_id", "comp_name"],
        "conflict_sql": "SELECT id FROM competitions WHERE student_id = ? AND comp_name = ?",
    },
    "企业实习": {
        "table": "internships", "columns": ["student_id", "company", "position", "description", "period"],
        "headers": ["学号", "公司", "职位", "描述", "时间"], "mode": "insert",
        "conflict_keys": ["student_id", "company", "period"],
        "conflict_sql": "SELECT id FROM internships WHERE student_id = ? AND company = ? AND period = ?",
    },
    "项目经历": {
        "table": "projects", "columns": ["student_id", "project_name", "rank", "description", "period"],
        "headers": ["学号", "项目名称", "排名", "描述", "时间"], "mode": "insert",
        "conflict_keys": ["student_id", "project_name"],
        "conflict_sql": "SELECT id FROM projects WHERE student_id = ? AND project_name = ?",
    },
}


@app.post("/api/admin/import/all")
async def admin_import_all(file: UploadFile = File(...), request: Request = None):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    file_name = file.filename or "unknown"
    if not file.filename.endswith((".xlsx", ".xls")):
        _log_import_error("格式错误", "仅支持 .xlsx/.xls", file_name)
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 或 .xls 格式")
    try:
        content = await file.read()
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as e:
        _log_import_error("文件解析失败", str(e), file_name)
        raise HTTPException(status_code=400, detail=f"无法解析 Excel: {str(e)}")
    conn = get_db()
    results = {}
    try:
        for sheet_name, config in SHEET_CONFIG.items():
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2: continue
            header = [str(h).strip() if h else "" for h in rows[0]]
            col_map = {}
            for i, h in enumerate(header):
                if h in config["headers"]: col_map[h] = i
            missing = [h for h in config["headers"] if h not in col_map]
            if missing:
                results[sheet_name] = {"success": 0, "error": f"缺少列: {', '.join(missing)}"}
                continue
            success, errors, batch = 0, [], 0
            for row_num, row in enumerate(rows[1:], start=2):
                if not row or all(c is None or str(c).strip() == "" for c in row): continue
                try:
                    vals = {}
                    for h, idx in col_map.items():
                        val = row[idx] if idx < len(row) else None
                        vals[h] = str(val).strip() if val is not None else ""
                    if config["mode"] == "upsert":
                        sid = vals["学号"]
                        pwd = vals.get("密码", "") or ""
                        gpa = float(vals.get("GPA", 0) or 0)
                        existing = conn.execute("SELECT id FROM students WHERE student_id = ?", (sid,)).fetchone()
                        if existing:
                            # 更新时不修改密码，除非明确提供了新密码
                            if pwd and len(pwd) >= 6:
                                pwd_hash = bcrypt.hashpw(pwd.encode(), bcrypt.gensalt()).decode()
                                conn.execute("UPDATE students SET name=?,college=?,major=?,class_name=?,gpa=?,password_hash=? WHERE student_id=?",
                                    (vals["姓名"], vals["学院"], vals["专业"], vals["班级"], gpa, pwd_hash, sid))
                            else:
                                conn.execute("UPDATE students SET name=?,college=?,major=?,class_name=?,gpa=? WHERE student_id=?",
                                    (vals["姓名"], vals["学院"], vals["专业"], vals["班级"], gpa, sid))
                        else:
                            pwd_hash = bcrypt.hashpw((pwd if pwd and len(pwd) >= 6 else "student123").encode(), bcrypt.gensalt()).decode()
                            conn.execute("INSERT INTO students (student_id,name,college,major,class_name,gpa,password_hash,role) VALUES (?,?,?,?,?,?,?,'student')",
                                (sid, vals["姓名"], vals["学院"], vals["专业"], vals["班级"], gpa, pwd_hash))
                    else:
                        conflict_sql = config.get("conflict_sql")
                        if conflict_sql:
                            ck = config["conflict_keys"]
                            check_params = [float(vals.get(h, 0)) if h in ("成绩","学分","实践学分","排名") else vals.get(h, "") for h in config["headers"] if h in ck]
                            if conn.execute(conflict_sql, check_params).fetchone():
                                row_data = {h: vals[h] for h in config["headers"]}
                                errors.append({"row": row_num, "data": row_data, "reason": "数据与已有记录冲突（同学生+同名称）"})
                                continue
                        cols = list(config["columns"])
                        params = []
                        for h, c in zip(config["headers"], cols):
                            v = vals[h]
                            if c in ("score", "credit", "practical_credit", "gpa"): v = float(v) if v else 0
                            elif c == "rank": v = int(v) if v else 1
                            params.append(v)
                        conn.execute(f"INSERT INTO {config['table']} ({','.join(cols)}) VALUES ({','.join(['?']*len(cols))})", params)
                    success += 1
                    batch += 1
                    if batch >= 500:  # 每500条提交一次，释放内存和锁
                        conn.commit()
                        batch = 0
                except Exception as e:
                    row_data = {}
                    try: row_data = {h: vals.get(h, "") for h in config["headers"]}
                    except: pass
                    errors.append({"row": row_num, "data": row_data, "reason": str(e)})
            if batch > 0:
                conn.commit()
            results[sheet_name] = {
                "success": success, "errors": errors[:50],
                "conflicts": sum(1 for e in errors if isinstance(e, dict) and "冲突" in e.get("reason", "")),
                "total_rows": len(rows) - 1,  # 总行数（含已处理+未处理）
                "processed": success + len([e for e in errors if isinstance(e, dict)]),  # 已处理行数
            }
            if errors:
                try:
                    saved = _save_conflicts_to_disk(sheet_name, errors)
                    results[sheet_name]["saved_to"] = saved
                except Exception: pass
        total = sum(r.get("success", 0) for r in results.values())
        return {"message": f"综合导入完成，共导入 {total} 条记录", "results": results}
    except Exception as e:
        _log_import_error("导入处理异常", str(e), file_name)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")
    finally:
        conn.close()


@app.post("/api/admin/import/conflicts/download")
async def admin_download_conflicts(request: Request):
    """下载冲突数据——按学号分文件，打包为 ZIP"""
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    body = await request.json()
    errors = body.get("errors", [])
    sheet_name = body.get("sheet", "未知")

    # 按学号分组
    groups = {}
    for e in errors:
        if isinstance(e, dict) and "data" in e:
            sid = e["data"].get("学号", "unknown")
            if sid not in groups:
                groups[sid] = []
            groups[sid].append(e)

    # 创建 ZIP
    import zipfile, os, time
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        if groups:
            headers = list(errors[0]["data"].keys()) if errors else []
            for sid, items in sorted(groups.items()):
                csv_buf = io.StringIO()
                writer = csv.writer(csv_buf)
                writer.writerow(["行号", "失败原因"] + headers)
                for e in items:
                    writer.writerow([e.get("row", ""), e.get("reason", "")] + [e["data"].get(h, "") for h in headers])
                zf.writestr(f"{sid}.csv", csv_buf.getvalue().encode("utf-8-sig"))

            # 汇总文件
            sum_buf = io.StringIO()
            writer = csv.writer(sum_buf)
            writer.writerow(["学号", "冲突条数"])
            for sid, items in sorted(groups.items()):
                writer.writerow([sid, len(items)])
            zf.writestr("_汇总.csv", sum_buf.getvalue().encode("utf-8-sig"))
        else:
            zf.writestr("无冲突数据.txt", "No conflicts found.".encode("utf-8"))

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    zip_buf.seek(0)
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=conflicts_{sheet_name}_{timestamp}.zip"},
    )


@app.get("/api/admin/template/all")
def admin_download_all_template(request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    import openpyxl
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    examples = {
        "学生基本信息": [["2021001", "张三", "计算机学院", "软件工程", "软工2101", "3.82", "student123"]],
        "课程成绩": [["2021001", "程序设计基础", "95", "4", "1.0", "2021秋", "学科基础"], ["2021001", "数据结构", "92", "4", "0.5", "2022春", "学科基础"]],
        "竞赛获奖": [["2021001", "全国大学生数学建模竞赛", "国家级", "二等奖", "2", "2022-11"]],
        "企业实习": [["2021001", "阿里巴巴", "后端开发实习生", "参与电商平台订单系统开发", "2023-07 —2023-09"]],
        "项目经历": [["2021001", "校园二手交易平台", "1", "基于Spring Boot搭建的校园二手交易系统", "2023-03 —2023-06"]],
    }
    first = True
    for sheet_name, config in SHEET_CONFIG.items():
        ws = wb.create_sheet(sheet_name, 0) if first else wb.create_sheet(sheet_name)
        first = False
        ws.append(config["headers"])
        for row in examples.get(sheet_name, []): ws.append(row)
        for i, h in enumerate(config["headers"], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(len(h) * 2, 14)
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_data_template.xlsx"})


# ========== 批量导入（CSV） ==========

IMPORT_TEMPLATES = {
    "courses":      ["学号", "课程名称", "成绩", "学分", "实践学分", "学期", "类别"],
    "competitions": ["学号", "竞赛名称", "级别", "获奖", "排名", "时间"],
    "internships":  ["学号", "公司", "职位", "描述", "时间"],
    "projects":     ["学号", "项目名称", "排名", "描述", "时间"],
}

IMPORT_SQL = {
    "courses":      "INSERT INTO courses (student_id, course_name, score, credit, practical_credit, semester, course_nature) VALUES (?, ?, ?, ?, ?, ?, ?)",
    "competitions": "INSERT INTO competitions (student_id, comp_name, level, award, rank, date) VALUES (?, ?, ?, ?, ?, ?)",
    "internships":  "INSERT INTO internships (student_id, company, position, description, period) VALUES (?, ?, ?, ?, ?)",
    "projects":     "INSERT INTO projects (student_id, project_name, rank, description, period) VALUES (?, ?, ?, ?, ?)",
}


@app.post("/api/admin/import/{data_type}")
async def admin_import(data_type: str, file: UploadFile = File(...), request: Request = None):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)
    file_name = file.filename or "unknown"

    if data_type not in IMPORT_TEMPLATES:
        _log_import_error("无效类型", f"不支持的数据类型: {data_type}", file_name)
        raise HTTPException(status_code=400, detail=f"不支持的数据类型: {data_type}")

    try:
        content = await file.read()
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("gbk")
    except Exception as e:
        _log_import_error("文件读取失败", str(e), file_name)
        raise HTTPException(status_code=400, detail=f"文件读取失败: {str(e)}")

    reader = csv.reader(io.StringIO(text))
    header = next(reader, None)
    if not header:
        raise HTTPException(status_code=400, detail="CSV 文件为空")

    # 去除表头空格
    header = [h.strip() for h in header]
    expected = IMPORT_TEMPLATES[data_type]

    # 检查必要列
    missing = [c for c in expected if c not in header]
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必要列: {', '.join(missing)}，模板列: {', '.join(expected)}")

    conn = get_db()
    try:
        success = 0
        errors = []
        for row_num, row in enumerate(reader, start=2):
            if not row or all(c.strip() == "" for c in row):
                continue
            try:
                record = {header[i]: row[i].strip() if i < len(row) else "" for i in range(len(header))}

                if data_type == "courses":
                    conn.execute(IMPORT_SQL["courses"], (
                        record["学号"], record["课程名称"],
                        float(record["成绩"]), float(record["学分"]),
                        float(record.get("实践学分", 0) or 0),
                        record["学期"], record["类别"],
                    ))
                elif data_type == "competitions":
                    conn.execute(IMPORT_SQL["competitions"], (
                        record["学号"], record["竞赛名称"],
                        record["级别"], record["获奖"],
                        int(record["排名"]), record["时间"],
                    ))
                elif data_type == "internships":
                    conn.execute(IMPORT_SQL["internships"], (
                        record["学号"], record["公司"],
                        record["职位"], record["描述"], record["时间"],
                    ))
                elif data_type == "projects":
                    conn.execute(IMPORT_SQL["projects"], (
                        record["学号"], record["项目名称"],
                        int(record["排名"]), record["描述"], record["时间"],
                    ))
                success += 1
            except Exception as e:
                try:
                    row_vals = {header[i]: row[i].strip() if i < len(row) else "" for i in range(len(header))}
                except:
                    row_vals = {}
                errors.append({"row": row_num, "data": row_vals, "reason": str(e)})

        conn.commit()
        if errors:
            try: _save_conflicts_to_disk(data_type, errors)
            except: pass
        return {
            "message": f"导入完成",
            "success": success,
            "errors": errors[:20],
        }
    except Exception as e:
        _log_import_error("CSV导入异常", str(e), file_name)
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")
    finally:
        conn.close()


@app.get("/api/admin/template/{data_type}")
def admin_download_template(data_type: str, request: Request):
    payload = verify_token(request.headers.get("Authorization", "").replace("Bearer ", ""))
    require_admin(payload)

    if data_type not in IMPORT_TEMPLATES:
        raise HTTPException(status_code=400, detail=f"不支持的数据类型: {data_type}")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(IMPORT_TEMPLATES[data_type])

    # 写一行示例数据
    examples = {
        "courses":      ["2021001", "程序设计基础", "95", "4", "1.0", "2021秋", "学科基础"],
        "competitions": ["2021001", "全国大学生数学建模竞赛", "国家级", "二等奖", "2", "2022-11"],
        "internships":  ["2021001", "阿里巴巴", "后端开发实习生", "参与电商平台订单系统开发", "2023-07 —2023-09"],
        "projects":     ["2021001", "校园二手交易平台", "1", "基于Spring Boot搭建的校园二手交易系统", "2023-03 —2023-06"],
    }
    writer.writerow(examples.get(data_type, []))

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={data_type}_template.csv"},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
